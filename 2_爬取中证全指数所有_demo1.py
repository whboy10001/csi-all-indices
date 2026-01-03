from playwright.sync_api import sync_playwright
import json
import time
import pandas as pd
import datetime

from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles.alignment import Alignment

# Excel导出功能
def export_to_excel(data: pd.DataFrame | list | None = None, filename: str = None) -> str:
    """将结构化的指数数据导出到Excel文件"""
    print('=== 数据导出到Excel ===\n')
    
    try:
        # 如果没有传入数据，从文件中读取
        if data is None:
            with open('csindex_structured_data.json', 'r', encoding='utf-8') as f:
                data = json.load(f)
        
        # 处理不同类型的数据
        data_length = 0
        if isinstance(data, pd.DataFrame):
            df = data.copy()
            data_length = len(df)
            print(f'共读取到 {data_length} 条指数数据')
        else:
            # 确保数据是列表格式
            if not isinstance(data, list):
                print('传入的数据不是列表或DataFrame格式')
                return ""
            
            data_length = len(data)
            print(f'共读取到 {data_length} 条指数数据')
            
            if not data:
                print('没有数据可导出')
                return ""
            
            # 将列表转换为DataFrame
            df = pd.DataFrame(data)
        
        # 生成文件名
        if not filename:
            today = datetime.datetime.now().strftime("%Y-%m-%d")
            filename = f"2_中证指数有限公司_指数列表_{today}.xlsx"
        
        # 导出DataFrame到Excel
        df.to_excel(filename, index=False, engine='openpyxl')
        
        # 加载工作簿和工作表
        wb = load_workbook(filename)
        ws = wb.active
        
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            # 计算列中最长字符串的宽度，考虑中文字符
            for cell in column:
                try:
                    if cell.value:
                        # 中文字符宽度计算为2，英文字符为1
                        text = str(cell.value)
                        width = 0
                        for char in text:
                            if ord(char) > 127:  # 中文字符
                                width += 2
                            else:  # 英文字符
                                width += 1
                        if width > max_length:
                            max_length = width
                except:
                    pass
            
            # 设置列宽，考虑Excel的列宽单位换算和余量
            # Excel的列宽单位约为0.83个字符宽度，加上适当余量
            adjusted_width = min(max_length * 1.0 + 3, 80)  # 增加最大宽度到80
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 设置所有列（除了首行）左对齐
        for row in ws.iter_rows(min_row=2):  # 从第2行开始
            for cell in row:
                cell.alignment = Alignment(horizontal='left')
        
        # 保存工作簿
        wb.save(filename)
        print(f'\n✅ 数据成功导出到 {filename}')
        print(f'📊 共导出 {data_length} 条指数数据')
        print(f'📋 包含 {len(df.columns)} 个字段')
        
        return filename
    except Exception as e:
        print(f'❌ 导出失败: {e}')
        import traceback
        traceback.print_exc()
        return ""

def crawl_all_csindex_data() -> pd.DataFrame:
    """爬取中证指数网站上的所有指数数据，支持分页"""
    print('=== 中证指数网站全量爬虫 ===\n')
    
    try:
        with sync_playwright() as p:
            # 启动浏览器
            print('正在启动浏览器...')
            browser = p.chromium.launch(
                headless=True,
                args=[
                    '--no-sandbox',
                    '--disable-dev-shm-usage'
                ]
            )
            
            # 创建上下文
            context = browser.new_context(
                user_agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
            )
            
            # 创建页面
            page = context.new_page()
            
            # 访问目标网址
            print('正在访问目标网址...')
            url = 'https://www.csindex.com.cn/#/indices/family/list'
            page.goto(url, wait_until='networkidle', timeout=60000)
            
            # 等待页面加载完成
            print('等待页面加载完成...')
            time.sleep(3)
            
            # 定义数据存储列表
            all_index_data = []
            
            # 定义数据列名
            columns = ['指数代码名称', '样本数量', '最新收盘', '近1个月收益率(%)', '资产类别', '指数热点', '指数币种', '合作指数', '指数类别', '发布时间']
            
            print('\n=== 开始分页爬取数据 ===')
            
            # 循环处理分页，获取所有288页数据
            page_num = 1
            max_pages = 288  # 用户提到的总页数
            
            while page_num <= max_pages:
                print(f'\n=== 处理第 {page_num} 页 ===')
                
                # 提取当前页的表格数据
                current_page_data = page.evaluate('''() => {
                    const tableData = [];
                    // 查找iview表格
                    const table = document.querySelector('.ivu-table');
                    if (table) {
                        const rows = table.querySelectorAll('.ivu-table-body tr');
                        for (const row of rows) {
                            const cells = row.querySelectorAll('td');
                            const rowData = [];
                            for (const cell of cells) {
                                rowData.push(cell.textContent.trim());
                            }
                            if (rowData.length > 5) {  // 过滤掉行数少于6的表格行
                                tableData.push(rowData);
                            }
                        }
                    }
                    return tableData;
                }''')
                
                if current_page_data:
                    print(f'  提取到 {len(current_page_data)} 条数据')
                    
                    # 检查是否为重复数据
                    if all_index_data and current_page_data == all_index_data[-len(current_page_data):]:
                        print(f'  检测到重复数据，停止爬取')
                        break
                    
                    all_index_data.extend(current_page_data)
                else:
                    print(f'  未提取到任何数据')
                
                # 尝试多种方式点击下一页
                next_page_found = False
                
                # 方式1: 使用iview分页组件的下一页按钮选择器
                try:
                    print('  方式1: 尝试点击iview下一页按钮...')
                    next_button = page.locator('.ivu-page-next')
                    
                    if next_button.is_visible() and next_button.is_enabled():
                        print('  点击下一页按钮')
                        next_button.click()
                        # 等待页面加载完成
                        time.sleep(5)
                        next_page_found = True
                        page_num += 1
                    else:
                        print('  iview下一页按钮不可见或不可用')
                except Exception as e:
                    print(f'  方式1失败: {e}')
                
                # 方式2: 使用更通用的选择器
                if not next_page_found:
                    try:
                        print('  方式2: 尝试点击文本为"下一页"的按钮...')
                        next_button = page.locator('button:has-text("下一页")')
                        
                        if next_button.is_visible() and next_button.is_enabled():
                            print('  点击下一页按钮')
                            next_button.click()
                            time.sleep(5)
                            next_page_found = True
                            page_num += 1
                        else:
                            print('  未找到文本为"下一页"的按钮')
                    except Exception as e:
                        print(f'  方式2失败: {e}')
                
                # 方式3: 尝试查找iview页码数字按钮，点击下一个页码
                if not next_page_found:
                    try:
                        print('  方式3: 尝试点击iview数字页码按钮...')
                        # 查找所有iview页码按钮
                        page_buttons = page.locator('.ivu-page-item')
                        button_count = page_buttons.count()
                        print(f'  找到 {button_count} 个页码按钮')
                        
                        # 点击当前页码的下一个按钮
                        if button_count > page_num:
                            print(f'  点击第 {page_num + 1} 个页码按钮')
                            page_buttons.nth(page_num).click()
                            time.sleep(5)
                            next_page_found = True
                            page_num += 1
                    except Exception as e:
                        print(f'  方式3失败: {e}')
                
                # 方式4: 尝试使用JavaScript直接修改iview页码
                if not next_page_found:
                    try:
                        print('  方式4: 尝试使用JavaScript修改iview页码...')
                        result = page.evaluate(f'''() => {{
                            // 查找iview页码输入框
                            const input = document.querySelector('.ivu-page-options input');
                            if (input) {{
                                input.value = {page_num + 1};
                                // 触发输入事件
                                input.dispatchEvent(new Event('input', {{ bubbles: true }}));
                                // 触发按键事件
                                input.dispatchEvent(new KeyboardEvent('keydown', {{ 
                                    bubbles: true, 
                                    key: 'Enter',
                                    code: 'Enter'
                                }}));
                                return true;
                            }}
                            return false;
                        }}''')
                        
                        if result:
                            print('  成功修改页码')
                            time.sleep(5)
                            next_page_found = True
                            page_num += 1
                        else:
                            print('  未找到iview页码输入框')
                    except Exception as e:
                        print(f'  方式4失败: {e}')
                
                # 方式5: 尝试模拟滚动加载
                if not next_page_found:
                    try:
                        print('  方式5: 尝试滚动加载更多数据...')
                        # 滚动到页面底部
                        page.evaluate('window.scrollTo(0, document.body.scrollHeight)')
                        time.sleep(3)
                        # 滚动回顶部
                        page.evaluate('window.scrollTo(0, 0)')
                        time.sleep(1)
                        # 再次滚动到底部
                        page.evaluate('window.scrollTo(0, document.body.scrollHeight)')
                        time.sleep(5)
                        next_page_found = True
                        page_num += 1
                    except Exception as e:
                        print(f'  方式5失败: {e}')
                
                # 如果所有方式都失败，停止爬取
                if not next_page_found:
                    print('  所有分页方式都失败，停止爬取')
                    break
            
            # 爬取完成，准备保存数据
            
            # 关闭浏览器
            browser.close()
            
            # 保存所有数据到文件
            if all_index_data:
                print(f'\n=== 爬取完成 ===')
                print(f'共提取到 {len(all_index_data)} 条指数数据')
                
                # 保存原始数据
                with open('csindex_raw_data.json', 'w', encoding='utf-8') as f:
                    json.dump(all_index_data, f, ensure_ascii=False, indent=2)
                print('原始数据已保存到 csindex_raw_data.json')
                
                # 保存带列名的结构化数据
                structured_data = []
                for row in all_index_data:
                    if len(row) == len(columns):
                        data_dict = dict(zip(columns, row))
                        # 拆分指数代码和名称
                        code_name = data_dict['指数代码名称']
                        import re
                        match = re.match(r'([0-9]{6,})(.*)', code_name)
                        if match:
                            data_dict['指数代码'] = match.group(1)
                            data_dict['指数名称'] = match.group(2).strip()
                        structured_data.append(data_dict)
                
                # 去重
                seen = set()
                unique_structured_data = []
                for item in structured_data:
                    key = item['指数代码名称']
                    if key not in seen:
                        seen.add(key)
                        unique_structured_data.append(item)
                
                print(f'去重后剩余 {len(unique_structured_data)} 条数据')
                
                with open('csindex_structured_data.json', 'w', encoding='utf-8') as f:
                    json.dump(unique_structured_data, f, ensure_ascii=False, indent=2)
                print('结构化数据已保存到 csindex_structured_data.json')
                
                # 直接导出到Excel
                print('\n直接导出数据到Excel...')
                export_to_excel(unique_structured_data)
                
                # 将结构化数据转换为DataFrame并返回
                return pd.DataFrame(unique_structured_data)
            else:
                print('\n=== 爬取完成 ===')
                print('未提取到任何指数数据')
                return pd.DataFrame()  # 返回空DataFrame
            
        return pd.DataFrame()  # 返回空DataFrame
    except Exception as e:
        print(f'✗ 爬取失败: {e}')
        import traceback
        traceback.print_exc()
        return pd.DataFrame()  # 异常情况下返回空DataFrame

if __name__ == '__main__':
    crawl_all_csindex_data()
