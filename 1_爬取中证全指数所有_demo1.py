#!/usr/bin/env python
# -*- coding:utf-8 -*-
"""
Date: 2025/8/4 14:00
Desc: 中证指数网站-指数列表
网站：https://www.csindex.com.cn/#/indices/family/list?index_series=1
"""
import warnings
from io import BytesIO
import datetime

import pandas as pd
import requests


def crawl_csindex_data() -> pd.DataFrame:
    """
    爬取中证指数网站-指数列表数据
    https://www.csindex.com.cn/#/indices/family/list?index_series=1
    Note: 但是不知道数据更新时间
    :return: 最新指数的列表,
    :rtype: pandas.DataFrame
    """
    url = (
        f"https://www.csindex.com.cn/csindex-home/exportExcel/indexAll/CH"
    )

    headers = {
        "Content-Type": "application/json;charset=UTF-8",
    }
    playloads = {
        "sorter": {
            "sortField": "null",
            "sortOrder": None
        },
        "pager": {
            "pageNum": 1,
            "pageSize": 10
        },
        "indexFilter": {
            "ifCustomized": None,
            "ifTracked": None,
            "ifWeightCapped": None,
            "indexCompliance": None,
            "hotSpot": None,
            "indexClassify": None,
            "currency": None,
            "region": None,
            "indexSeries": ["1"],
            "undefined": None
        }
    }
    r = requests.post(url, json=playloads, headers=headers)

    temp_df = pd.read_excel(BytesIO(r.content))
    temp_df["基日"] = pd.to_datetime(
        temp_df["基日"], format="%Y-%m-%d", errors="coerce"
    ).dt.date
    temp_df["发布时间"] = pd.to_datetime(
        temp_df["发布时间"], format="%Y-%m-%d", errors="coerce"
    ).dt.date
    temp_df["指数代码"] = temp_df["指数代码"].astype(str).str.zfill(6)
    
    return temp_df


def export_to_excel(dataframe: pd.DataFrame, filename: str = None) -> str:
    """
    将DataFrame导出为Excel文件，并自动设置列宽和对齐方式
    :param dataframe: 要导出的数据
    :type dataframe: pd.DataFrame
    :param filename: 导出文件名，不指定则自动生成带日期的文件名
    :type filename: str, optional
    :return: 导出的文件名
    :rtype: str
    """
    from openpyxl.utils import get_column_letter
    from openpyxl import load_workbook
    from openpyxl.styles.alignment import Alignment
    
    # 生成文件名
    if not filename:
        today = datetime.datetime.now().strftime("%Y-%m-%d")
        filename = f"1_中证指数列表_{today}.xlsx"
    
    # 导出DataFrame到Excel
    dataframe.to_excel(filename, index=False, engine='openpyxl')
    
    # 加载工作簿和工作表
    wb = load_workbook(filename)
    ws = wb.active
    
    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        # 计算列中最长字符串的宽度，考虑中文字符
        for cell in column:
            try:
                if cell.value:
                    # 中文字符宽度计算为2，英文字符为1
                    text = str(cell.value)
                    width = 0
                    for char in text:
                        if ord(char) > 127:  # 中文字符
                            width += 2
                        else:  # 英文字符
                            width += 1
                    if width > max_length:
                        max_length = width
            except:
                pass
        
        # 设置列宽，考虑Excel的列宽单位换算和余量
        # Excel的列宽单位约为0.83个字符宽度，加上适当余量
        adjusted_width = min(max_length * 1.0 + 3, 80)  # 增加最大宽度到80
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 设置所有列（除了首行）左对齐
    for row in ws.iter_rows(min_row=2):  # 从第2行开始
        for cell in row:
            cell.alignment = Alignment(horizontal='left')
    
    # 保存工作簿
    wb.save(filename)
    print(f"数据已成功导出到: {filename}")
    
    return filename


def index_csindex_all(export_excel: bool = True) -> pd.DataFrame:
    """
    中证指数网站-指数列表主函数
    https://www.csindex.com.cn/#/indices/family/list?index_series=1
    :param export_excel: 是否导出为Excel文件, defaults to True
    :type export_excel: bool, optional
    :return: 最新指数的列表,
    :rtype: pandas.DataFrame
    """
    # 爬取数据
    df = crawl_csindex_data()
    
    # 导出Excel
    if export_excel:
        export_to_excel(df)
    
    return df


if __name__ == "__main__":
    index_csindex_all_df = index_csindex_all()
    print(index_csindex_all_df)
